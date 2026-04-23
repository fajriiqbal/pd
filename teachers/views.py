from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render

from academics.models import ClassSubject

from .forms import (
    TeacherAdditionalTaskForm,
    TeacherEducationHistoryForm,
    TeacherMutationRecordForm,
    TeacherImportUploadForm,
    TeacherRecordForm,
    TeacherTeachingAssignmentForm,
)
from .import_utils import (
    TEACHER_IMPORT_HEADERS,
    build_teacher_import_preview,
    delete_import_preview,
    execute_teacher_import,
    load_import_preview,
    save_import_preview,
)
from .reference import search_school_reference
from .models import TeacherAdditionalTask, TeacherArchive, TeacherEducationHistory, TeacherMutationRecord, TeacherProfile


@login_required
def teacher_list(request):
    query = request.GET.get("q", "").strip()
    teachers = TeacherProfile.objects.select_related("user").annotate(
        active_teaching_count=Count("class_subjects", filter=Q(class_subjects__is_active=True), distinct=True),
        active_task_count=Count("additional_tasks", filter=Q(additional_tasks__is_active=True), distinct=True),
    )

    if query:
        teachers = teachers.filter(
            Q(user__full_name__icontains=query)
            | Q(user__username__icontains=query)
            | Q(nip__icontains=query)
            | Q(subject__icontains=query)
            | Q(class_subjects__subject__name__icontains=query)
            | Q(additional_tasks__name__icontains=query)
        ).distinct()

    import_result = request.session.pop("teacher_import_result", None)
    context = {
        "teachers": teachers,
        "query": query,
        "import_result": import_result,
        "active_teacher_count": TeacherProfile.objects.filter(is_active=True).count(),
        "assigned_teacher_count": TeacherProfile.objects.filter(class_subjects__is_active=True).distinct().count(),
        "active_task_count": TeacherAdditionalTask.objects.filter(is_active=True).count(),
    }
    return render(request, "teachers/teacher_list.html", context)


@login_required
def teaching_assignment_list(request):
    query = request.GET.get("q", "").strip()
    assignments = ClassSubject.objects.select_related(
        "school_class",
        "subject",
        "teacher__user",
    ).order_by("teacher__user__full_name", "school_class__level_order", "subject__sort_order")

    if query:
        assignments = assignments.filter(
            Q(teacher__user__full_name__icontains=query)
            | Q(subject__name__icontains=query)
            | Q(subject__code__icontains=query)
            | Q(school_class__name__icontains=query)
        )

    active_assignments = assignments.filter(is_active=True)
    context = {
        "assignments": assignments,
        "query": query,
        "assignment_count": assignments.count(),
        "active_assignment_count": active_assignments.count(),
        "teacher_with_assignment_count": active_assignments.exclude(teacher__isnull=True).values("teacher").distinct().count(),
        "total_weekly_hours": active_assignments.aggregate(total=Sum("weekly_hours"))["total"] or 0,
    }
    return render(request, "teachers/teaching_assignment_list.html", context)


@login_required
def teaching_assignment_create(request):
    form = TeacherTeachingAssignmentForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Pengaturan mapel guru berhasil ditambahkan.")
        return redirect("teachers:teaching_assignments")

    return render(
        request,
        "shared/form_page.html",
        {
            "form": form,
            "page_kicker": "Beban Mengajar",
            "page_title": "Tambah mapel yang diajar",
            "page_description": "Hubungkan guru dengan mapel, kelas, KKM, dan jumlah jam pelajaran.",
            "submit_label": "Simpan pengaturan",
            "cancel_url": "teachers:teaching_assignments",
            "checkbox_fields": ["is_active"],
        },
    )


@login_required
def teaching_assignment_update(request, pk):
    assignment = get_object_or_404(ClassSubject, pk=pk)
    form = TeacherTeachingAssignmentForm(request.POST or None, instance=assignment)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Pengaturan mapel guru berhasil diperbarui.")
        return redirect("teachers:teaching_assignments")

    return render(
        request,
        "shared/form_page.html",
        {
            "form": form,
            "page_kicker": "Beban Mengajar",
            "page_title": f"Edit mapel yang diajar {assignment}",
            "page_description": "Perbarui guru pengampu, KKM, jam pelajaran, atau status aktif.",
            "submit_label": "Update pengaturan",
            "cancel_url": "teachers:teaching_assignments",
            "checkbox_fields": ["is_active"],
        },
    )


@login_required
def teaching_assignment_delete(request, pk):
    assignment = get_object_or_404(ClassSubject, pk=pk)
    if assignment.grade_books.exists():
        messages.error(request, "Pengaturan ini tidak bisa dihapus karena sudah memiliki ledger nilai.")
        return redirect("teachers:teaching_assignments")

    if request.method == "POST":
        assignment.delete()
        messages.success(request, "Pengaturan mapel guru berhasil dihapus.")
        return redirect("teachers:teaching_assignments")

    return render(
        request,
        "shared/confirm_delete.html",
        {
            "item_name": str(assignment),
            "item_type": "mapel yang diajar",
            "cancel_url": "teachers:teaching_assignments",
        },
    )


@login_required
def additional_task_list(request):
    query = request.GET.get("q", "").strip()
    tasks = TeacherAdditionalTask.objects.select_related("teacher__user").order_by(
        "teacher__user__full_name",
        "task_type",
        "name",
    )

    if query:
        tasks = tasks.filter(
            Q(teacher__user__full_name__icontains=query)
            | Q(name__icontains=query)
            | Q(description__icontains=query)
        )

    active_tasks = tasks.filter(is_active=True)
    context = {
        "tasks": tasks,
        "query": query,
        "task_count": tasks.count(),
        "active_task_count": active_tasks.count(),
        "teacher_with_task_count": active_tasks.values("teacher").distinct().count(),
        "total_task_hours": active_tasks.aggregate(total=Sum("hours_per_week"))["total"] or 0,
    }
    return render(request, "teachers/additional_task_list.html", context)


@login_required
def additional_task_create(request):
    form = TeacherAdditionalTaskForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Tugas tambahan guru berhasil ditambahkan.")
        return redirect("teachers:additional_tasks")

    return render(
        request,
        "shared/form_page.html",
        {
            "form": form,
            "page_kicker": "Tugas Tambahan",
            "page_title": "Tambah tugas tambahan",
            "page_description": "Catat amanah tambahan guru seperti pembina, koordinator, piket, atau tugas madrasah lain.",
            "submit_label": "Simpan tugas",
            "cancel_url": "teachers:additional_tasks",
            "checkbox_fields": ["is_active"],
        },
    )


@login_required
def additional_task_update(request, pk):
    task = get_object_or_404(TeacherAdditionalTask, pk=pk)
    form = TeacherAdditionalTaskForm(request.POST or None, instance=task)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Tugas tambahan guru berhasil diperbarui.")
        return redirect("teachers:additional_tasks")

    return render(
        request,
        "shared/form_page.html",
        {
            "form": form,
            "page_kicker": "Tugas Tambahan",
            "page_title": f"Edit tugas {task.name}",
            "page_description": "Perbarui penanggung jawab, jenis tugas, periode, atau status aktif.",
            "submit_label": "Update tugas",
            "cancel_url": "teachers:additional_tasks",
            "checkbox_fields": ["is_active"],
        },
    )


@login_required
def additional_task_delete(request, pk):
    task = get_object_or_404(TeacherAdditionalTask, pk=pk)
    if request.method == "POST":
        task.delete()
        messages.success(request, "Tugas tambahan guru berhasil dihapus.")
        return redirect("teachers:additional_tasks")

    return render(
        request,
        "shared/confirm_delete.html",
        {
            "item_name": str(task),
            "item_type": "tugas tambahan guru",
            "cancel_url": "teachers:additional_tasks",
        },
    )


def _sync_teacher_archive(teacher, mutation_record):
    archive, _ = TeacherArchive.objects.get_or_create(teacher=teacher)
    archive.full_name = teacher.user.full_name
    archive.nip = teacher.nip or ""
    archive.nik = teacher.nik or ""
    archive.nuptk = teacher.nuptk or ""
    archive.subject = teacher.subject
    archive.task = teacher.task
    archive.placement = teacher.placement
    archive.total_jtm = teacher.total_jtm
    archive.gender = teacher.gender
    archive.birth_place = teacher.birth_place
    archive.birth_date = teacher.birth_date
    archive.address = teacher.address
    archive.hire_date = teacher.hire_date
    archive.madrasah_email = teacher.madrasah_email
    archive.employment_status = teacher.employment_status
    archive.exit_status = mutation_record.exit_status or TeacherArchive.ExitStatus.OTHER
    archive.exit_notes = mutation_record.notes or mutation_record.reason
    archive.save()


@login_required
def teacher_mutation_list(request):
    query = request.GET.get("q", "").strip()
    direction = request.GET.get("direction", "").strip()
    mutations = TeacherMutationRecord.objects.select_related("teacher__user", "created_by").order_by("-mutation_date", "-created_at")

    if query:
        mutations = mutations.filter(
            Q(teacher__user__full_name__icontains=query)
            | Q(teacher__nip__icontains=query)
            | Q(origin_school_name__icontains=query)
            | Q(destination_school_name__icontains=query)
        )

    if direction in {TeacherMutationRecord.Direction.INBOUND, TeacherMutationRecord.Direction.OUTBOUND}:
        mutations = mutations.filter(direction=direction)

    context = {
        "mutations": mutations,
        "query": query,
        "direction": direction,
        "inbound_count": TeacherMutationRecord.objects.filter(direction=TeacherMutationRecord.Direction.INBOUND).count(),
        "outbound_count": TeacherMutationRecord.objects.filter(direction=TeacherMutationRecord.Direction.OUTBOUND).count(),
    }
    return render(request, "teachers/mutation_list.html", context)


@login_required
def teacher_mutation_create(request):
    form = TeacherMutationRecordForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        mutation = form.save(commit=False)
        mutation.created_by = request.user
        mutation.save()

        teacher = mutation.teacher
        if mutation.direction == TeacherMutationRecord.Direction.INBOUND:
            if mutation.destination_placement:
                teacher.placement = mutation.destination_placement
            teacher.is_active = True
            teacher.user.is_school_active = True
            teacher.user.save(update_fields=["is_school_active"])
            teacher.save(update_fields=["placement", "is_active", "updated_at"])
        else:
            teacher.is_active = False
            teacher.user.is_school_active = False
            teacher.user.save(update_fields=["is_school_active"])
            teacher.save(update_fields=["is_active", "updated_at"])
            _sync_teacher_archive(teacher, mutation)

        messages.success(request, "Data mutasi guru berhasil disimpan.")
        return redirect("teachers:edit", pk=teacher.pk)

    return render(
        request,
        "shared/form_page.html",
        {
            "form": form,
            "page_kicker": "Mutasi Guru",
            "page_title": "Tambah mutasi guru",
            "page_description": "Catat mutasi masuk atau keluar guru beserta status akhir seperti pensiun atau pindah.",
            "submit_label": "Simpan mutasi",
            "cancel_url": "teachers:list",
            "checkbox_fields": [],
        },
    )


@login_required
def teacher_archive_list(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    archives = TeacherArchive.objects.select_related("teacher__user").order_by("-archived_at", "full_name")

    if query:
        archives = archives.filter(
            Q(full_name__icontains=query)
            | Q(nip__icontains=query)
            | Q(nuptk__icontains=query)
            | Q(subject__icontains=query)
        )

    if status in {TeacherArchive.ExitStatus.PENSIONED, TeacherArchive.ExitStatus.TRANSFERRED, TeacherArchive.ExitStatus.OTHER}:
        archives = archives.filter(exit_status=status)

    context = {
        "archives": archives,
        "query": query,
        "status": status,
        "archive_count": TeacherArchive.objects.count(),
    }
    return render(request, "teachers/archive_list.html", context)


@login_required
def school_reference_search(request):
    query = request.GET.get("q", "").strip()
    results = search_school_reference(query)
    return JsonResponse({"query": query, "results": results})


@login_required
def teacher_import_preview(request):
    if request.method != "POST":
        return redirect("teachers:list")

    form = TeacherImportUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "File import guru belum valid. Periksa file dan password default.")
        return redirect("teachers:list")

    preview_payload = build_teacher_import_preview(
        form.cleaned_data["excel_file"],
        form.cleaned_data["default_password"],
    )

    if not preview_payload.get("ok"):
        request.session["teacher_import_result"] = {
            "title": "Preview import guru gagal",
            "created": 0,
            "updated": 0,
            "failed": len(preview_payload.get("errors", [])),
            "errors": preview_payload.get("errors", []),
        }
        return redirect("teachers:list")

    token = save_import_preview(preview_payload)
    return render(
        request,
        "teachers/import_preview.html",
        {
            "preview": preview_payload,
            "preview_token": token,
            "page_kicker": "Import Guru",
            "page_title": "Preview import data guru",
            "page_description": "Periksa dulu data guru yang akan dibuat, diupdate, atau gagal dibaca sebelum import dijalankan.",
        },
    )


@login_required
def teacher_import_execute(request):
    if request.method != "POST":
        return redirect("teachers:list")

    token = request.POST.get("preview_token", "").strip()
    preview_payload = load_import_preview(token)
    if not preview_payload:
        messages.error(request, "Data preview import guru tidak ditemukan atau sudah kedaluwarsa.")
        return redirect("teachers:list")

    result = execute_teacher_import(preview_payload)
    delete_import_preview(token)

    request.session["teacher_import_result"] = {
        "title": "Import guru selesai",
        "created": result["created"],
        "updated": result["updated"],
        "failed": result["failed"],
        "errors": result["errors"][:10],
    }
    return redirect("teachers:list")


@login_required
def teacher_import_template(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, "Paket openpyxl belum terpasang. Jalankan `pip install -r requirements.txt` terlebih dahulu.")
        return redirect("teachers:list")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data Guru"
    sheet.append(TEACHER_IMPORT_HEADERS)
    sheet.append(
        [
            "Contoh Guru",
            "3201010101010001",
            "1234567890123456",
            "Honorer",
            "",
            "L",
            "Bandung",
            "1988-05-12",
            "081234567890",
            "guru@example.com",
            "guru@madrasah.id",
            "guru12345",
            "Guru Mapel",
            "Fikih",
            "MTs",
            24,
        ]
    )
    sheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="DCFCE7")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="14532D")
        cell.fill = header_fill

    for column_idx, header in enumerate(TEACHER_IMPORT_HEADERS, start=1):
        width = max(len(header) + 4, 16)
        sheet.column_dimensions[get_column_letter(column_idx)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="template-import-guru.xlsx"'
    workbook.save(response)
    return response


@login_required
def teacher_create(request):
    form = TeacherRecordForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Data guru berhasil ditambahkan.")
        return redirect("teachers:list")

    return render(
        request,
        "shared/form_page.html",
        {
            "form": form,
            "page_kicker": "Input Guru",
            "page_title": "Tambah data guru",
            "page_description": "Lengkapi akun login dan profil guru dalam satu form operator.",
            "submit_label": "Simpan guru",
            "cancel_url": "teachers:list",
            "checkbox_fields": ["is_school_active", "is_active"],
        },
    )


@login_required
def teacher_update(request, pk):
    teacher = get_object_or_404(TeacherProfile.objects.select_related("user"), pk=pk)
    form = TeacherRecordForm(request.POST or None, instance=teacher)

    if request.method == "POST":
        form_type = request.POST.get("form_type", "data_diri")
        if form_type == "data_diri" and form.is_valid():
            form.save()
            messages.success(request, "Data diri guru berhasil diperbarui.")
            return redirect("teachers:edit", pk=teacher.pk)

    education_histories = teacher.education_histories.all()
    education_form = TeacherEducationHistoryForm()
    teaching_assignments = teacher.class_subjects.select_related(
        "school_class",
        "subject",
    ).order_by("school_class__level_order", "subject__sort_order", "subject__name")
    additional_tasks = teacher.additional_tasks.order_by("-is_active", "task_type", "name")
    active_teaching = teaching_assignments.filter(is_active=True)

    context = {
        "teacher": teacher,
        "form": form,
        "education_form": education_form,
        "education_degree_choices": TeacherEducationHistory.DegreeLevel.choices,
        "education_histories": education_histories,
        "teaching_assignments": teaching_assignments,
        "additional_tasks": additional_tasks,
        "teaching_assignment_count": active_teaching.count(),
        "teaching_hours": active_teaching.aggregate(total=Sum("weekly_hours"))["total"] or 0,
        "additional_task_count": additional_tasks.filter(is_active=True).count(),
        "active_additional_tasks": additional_tasks.filter(is_active=True),
        "checkbox_fields": ["is_school_active", "is_active"],
    }
    return render(request, "teachers/teacher_profile.html", context)


@login_required
def teacher_education_add(request, pk):
    teacher = get_object_or_404(TeacherProfile, pk=pk)
    if request.method != "POST":
        return redirect("teachers:edit", pk=teacher.pk)

    form = TeacherEducationHistoryForm(request.POST, request.FILES)
    if form.is_valid():
        education = form.save(commit=False)
        education.teacher = teacher
        education.save()
        if education.is_highest_degree:
            teacher.education_histories.exclude(pk=education.pk).update(is_highest_degree=False)
        messages.success(request, "Riwayat pendidikan berhasil ditambahkan.")
    else:
        messages.error(request, "Riwayat pendidikan gagal disimpan. Periksa form terlebih dahulu.")
    return redirect("teachers:edit", pk=teacher.pk)


@login_required
def teacher_education_update(request, pk, education_pk):
    teacher = get_object_or_404(TeacherProfile, pk=pk)
    education = get_object_or_404(TeacherEducationHistory, pk=education_pk, teacher=teacher)
    if request.method != "POST":
        return redirect("teachers:edit", pk=teacher.pk)

    form = TeacherEducationHistoryForm(request.POST, request.FILES, instance=education)
    if form.is_valid():
        education = form.save()
        if education.is_highest_degree:
            teacher.education_histories.exclude(pk=education.pk).update(is_highest_degree=False)
        messages.success(request, "Riwayat pendidikan berhasil diperbarui.")
    else:
        messages.error(request, "Perubahan pendidikan gagal disimpan. Data yang diisi belum valid.")
    return redirect("teachers:edit", pk=teacher.pk)


@login_required
def teacher_education_delete(request, pk, education_pk):
    teacher = get_object_or_404(TeacherProfile, pk=pk)
    education = get_object_or_404(TeacherEducationHistory, pk=education_pk, teacher=teacher)
    if request.method == "POST":
        education.delete()
        messages.success(request, "Riwayat pendidikan berhasil dihapus.")
    return redirect("teachers:edit", pk=teacher.pk)


@login_required
def teacher_delete(request, pk):
    teacher = get_object_or_404(TeacherProfile.objects.select_related("user"), pk=pk)
    if teacher.homeroom_groups.exists():
        messages.error(request, "Guru ini tidak bisa dihapus karena masih menjadi wali kelas.")
        return redirect("teachers:list")

    if request.method == "POST":
        with transaction.atomic():
            teacher.user.delete()
        messages.success(request, "Data guru berhasil dihapus.")
        return redirect("teachers:list")

    return render(
        request,
        "shared/confirm_delete.html",
        {
            "item_name": teacher.user.full_name,
            "item_type": "data guru",
            "cancel_url": "teachers:list",
        },
    )
