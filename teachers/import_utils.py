import json
import re
import uuid
from datetime import date, datetime
from pathlib import Path

from django.conf import settings
from django.db import transaction

from accounts.models import CustomUser

from .models import TeacherProfile

IMPORT_CACHE_DIR = Path(settings.MEDIA_ROOT) / "import_cache" / "teachers"

TEACHER_IMPORT_HEADERS = [
    "Nama Lengkap",
    "NIK",
    "NUPTK",
    "Status Kepegawaian",
    "NIP",
    "Jenis Kelamin",
    "Tempat Lahir",
    "Tanggal Lahir",
    "Nomor Handphone",
    "Email",
    "Email Akun Madrasah Digital",
    "Password Awal",
    "Tugas",
    "Mata Pelajaran",
    "Penempatan",
    "Total JTM",
]

HEADER_ALIASES = {
    "nama": "full_name",
    "nama lengkap": "full_name",
    "nama guru": "full_name",
    "nik": "nik",
    "nuptk": "nuptk",
    "status kepegawaian": "employment_status",
    "nip": "nip",
    "jenis kelamin": "gender",
    "jk": "gender",
    "tempat lahir": "birth_place",
    "tgl lahir": "birth_date",
    "tanggal lahir": "birth_date",
    "nomor handphone": "phone_number",
    "nomor hp": "phone_number",
    "no hp": "phone_number",
    "hp": "phone_number",
    "email": "email",
    "email akun madrasah digital": "madrasah_email",
    "akun madrasah digital": "madrasah_email",
    "password awal": "initial_password",
    "password": "initial_password",
    "tugas": "task",
    "mata pelajaran": "subject",
    "mapel": "subject",
    "penempatan": "placement",
    "total jtm": "total_jtm",
    "jtm": "total_jtm",
}


def _normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("\n", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def _normalize_number_text(value):
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).rstrip("0").rstrip(".")
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def _parse_date(value):
    if value in (None, ""):
        return None, None
    if isinstance(value, datetime):
        return value.date().isoformat(), None
    if isinstance(value, date):
        return value.isoformat(), None

    text = _clean_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat(), None
        except ValueError:
            continue
    return None, f"Tanggal lahir tidak valid: {text}"


def _parse_iso_date(value):
    if not value:
        return None
    if isinstance(value, date):
        return value
    return date.fromisoformat(value)


def _parse_gender(value):
    text = _clean_text(value).lower()
    if text in {"l", "lk", "laki laki", "laki-laki", "male"}:
        return TeacherProfile.Gender.MALE
    if text in {"p", "pr", "perempuan", "female"}:
        return TeacherProfile.Gender.FEMALE
    return ""


def _parse_employment_status(value):
    text = _clean_text(value).lower()
    if text in {"", "tetap", "pns", "gt", "guru tetap"}:
        return TeacherProfile.EmploymentStatus.PERMANENT, None
    if text in {"honorer", "gtt", "guru honorer", "non pns", "non-pns"}:
        return TeacherProfile.EmploymentStatus.HONORARY, None
    if text in {"kontrak", "guru kontrak", "contract"}:
        return TeacherProfile.EmploymentStatus.CONTRACT, None
    return None, f"Status kepegawaian tidak dikenali: {value}"


def _parse_total_jtm(value):
    if value in (None, ""):
        return 0, None
    text = _normalize_number_text(value)
    if not re.fullmatch(r"\d+", text):
        return None, f"Total JTM harus berupa angka: {value}"
    return int(text), None


def _generate_unique_username(preferred, exclude_user_id=None):
    base = _clean_text(preferred).lower().replace(" ", "")
    base = re.sub(r"[^a-z0-9.@_-]", "", base) or "guru"
    candidate = base[:150]
    counter = 1
    qs = CustomUser.objects.all()
    if exclude_user_id:
        qs = qs.exclude(pk=exclude_user_id)
    while qs.filter(username=candidate).exists():
        suffix = f"-{counter}"
        candidate = f"{base[:150 - len(suffix)]}{suffix}"
        counter += 1
    return candidate


def _cache_file_path(token):
    IMPORT_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    return IMPORT_CACHE_DIR / f"{token}.json"


def save_import_preview(payload):
    token = uuid.uuid4().hex
    _cache_file_path(token).write_text(json.dumps(payload, ensure_ascii=True), encoding="utf-8")
    return token


def load_import_preview(token):
    path = _cache_file_path(token)
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def delete_import_preview(token):
    path = _cache_file_path(token)
    if path.exists():
        path.unlink()


def _append_error(errors, sheet, row, message):
    errors.append({"sheet": sheet, "row": row, "message": message})


def _build_existing_indexes():
    teachers = list(TeacherProfile.objects.select_related("user").all())
    return {
        "nip": {teacher.nip: teacher for teacher in teachers if teacher.nip},
        "nik": {teacher.nik: teacher for teacher in teachers if teacher.nik},
        "nuptk": {teacher.nuptk: teacher for teacher in teachers if teacher.nuptk},
        "madrasah_email": {
            teacher.madrasah_email.lower(): teacher
            for teacher in teachers
            if teacher.madrasah_email
        },
        "username": {
            teacher.user.username.lower(): teacher
            for teacher in teachers
            if teacher.user_id and teacher.user.username
        },
    }


def _resolve_existing_teacher(row_data, indexes):
    matches = []
    identifiers = (
        ("nip", row_data["nip"]),
        ("nik", row_data["nik"]),
        ("nuptk", row_data["nuptk"]),
        ("madrasah_email", row_data["madrasah_email"].lower() if row_data["madrasah_email"] else ""),
        ("username", row_data["madrasah_email"].lower() if row_data["madrasah_email"] else ""),
    )

    for key, value in identifiers:
        if value and value in indexes[key]:
            matches.append(indexes[key][value])

    unique_ids = {teacher.id for teacher in matches}
    if len(unique_ids) > 1:
        return None, "NIP/NIK/NUPTK/Email akun mengarah ke beberapa guru berbeda."
    return matches[0] if matches else None, None


def build_teacher_import_preview(uploaded_file, default_password):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "ok": False,
            "errors": [
                {
                    "sheet": "-",
                    "row": "-",
                    "message": "Paket openpyxl belum terpasang. Jalankan `pip install -r requirements.txt` terlebih dahulu.",
                }
            ],
        }

    workbook = load_workbook(uploaded_file, data_only=True)
    indexes = _build_existing_indexes()

    rows_payload = []
    errors = []
    create_count = 0
    update_count = 0
    file_seen = set()

    for sheet in workbook.worksheets:
        if sheet.max_row < 1:
            _append_error(errors, sheet.title, "-", "Sheet kosong.")
            continue

        headers = {}
        for idx, cell in enumerate(sheet[1], start=1):
            normalized = _normalize_header(cell.value)
            if normalized in HEADER_ALIASES:
                headers[HEADER_ALIASES[normalized]] = idx

        if "full_name" not in headers:
            _append_error(errors, sheet.title, 1, "Header wajib tidak ditemukan: Nama Lengkap.")
            continue

        for row_idx in range(2, sheet.max_row + 1):
            row_values = {
                key: sheet.cell(row=row_idx, column=col_idx).value
                for key, col_idx in headers.items()
            }
            row_data = {
                "full_name": _clean_text(row_values.get("full_name")),
                "nik": _normalize_number_text(row_values.get("nik")),
                "nuptk": _normalize_number_text(row_values.get("nuptk")),
                "nip": _normalize_number_text(row_values.get("nip")),
                "birth_place": _clean_text(row_values.get("birth_place")),
                "phone_number": _normalize_number_text(row_values.get("phone_number")),
                "email": _clean_text(row_values.get("email")),
                "madrasah_email": _clean_text(row_values.get("madrasah_email")),
                "initial_password": _clean_text(row_values.get("initial_password")) or default_password,
                "task": _clean_text(row_values.get("task")),
                "subject": _clean_text(row_values.get("subject")),
                "placement": _clean_text(row_values.get("placement")),
            }

            if not any(row_data.values()) and not row_values.get("gender"):
                continue

            if not row_data["full_name"]:
                _append_error(errors, sheet.title, row_idx, "Nama lengkap kosong.")
                continue

            gender = _parse_gender(row_values.get("gender"))
            if not gender:
                _append_error(errors, sheet.title, row_idx, "Jenis kelamin tidak dikenali.")
                continue

            employment_status, status_error = _parse_employment_status(row_values.get("employment_status"))
            if status_error:
                _append_error(errors, sheet.title, row_idx, status_error)
                continue

            birth_date, birth_date_error = _parse_date(row_values.get("birth_date"))
            if birth_date_error:
                _append_error(errors, sheet.title, row_idx, birth_date_error)
                continue

            total_jtm, total_jtm_error = _parse_total_jtm(row_values.get("total_jtm"))
            if total_jtm_error:
                _append_error(errors, sheet.title, row_idx, total_jtm_error)
                continue

            duplicate_keys = [
                f"{key}:{value.lower() if key == 'madrasah_email' else value}"
                for key, value in (
                    ("nip", row_data["nip"]),
                    ("nik", row_data["nik"]),
                    ("nuptk", row_data["nuptk"]),
                    ("madrasah_email", row_data["madrasah_email"]),
                )
                if value
            ]
            duplicate_in_file = next((key for key in duplicate_keys if key in file_seen), None)
            if duplicate_in_file:
                _append_error(errors, sheet.title, row_idx, f"Data duplikat di file import: {duplicate_in_file}.")
                continue
            file_seen.update(duplicate_keys)

            existing_teacher, conflict_error = _resolve_existing_teacher(row_data, indexes)
            if conflict_error:
                _append_error(errors, sheet.title, row_idx, conflict_error)
                continue

            action = "update" if existing_teacher else "create"
            if action == "create":
                create_count += 1
            else:
                update_count += 1

            rows_payload.append(
                {
                    "sheet": sheet.title,
                    "row": row_idx,
                    "action": action,
                    "teacher_id": existing_teacher.id if existing_teacher else None,
                    **row_data,
                    "gender": gender,
                    "employment_status": employment_status,
                    "birth_date": birth_date,
                    "total_jtm": total_jtm,
                    "is_active": True,
                }
            )

    return {
        "ok": len(rows_payload) > 0 or len(errors) == 0,
        "default_password": default_password,
        "rows": rows_payload,
        "errors": errors,
        "summary": {
            "total_rows": len(rows_payload) + len(errors),
            "ready_count": len(rows_payload),
            "create_count": create_count,
            "update_count": update_count,
            "error_count": len(errors),
        },
    }


def execute_teacher_import(preview_payload):
    result = {"created": 0, "updated": 0, "failed": 0, "errors": []}

    for item in preview_payload["rows"]:
        try:
            with transaction.atomic():
                teacher = None
                if item["teacher_id"]:
                    teacher = TeacherProfile.objects.select_related("user").filter(pk=item["teacher_id"]).first()
                if not teacher and item["nip"]:
                    teacher = TeacherProfile.objects.select_related("user").filter(nip=item["nip"]).first()
                if not teacher and item["nuptk"]:
                    teacher = TeacherProfile.objects.select_related("user").filter(nuptk=item["nuptk"]).first()
                if not teacher and item["nik"]:
                    teacher = TeacherProfile.objects.select_related("user").filter(nik=item["nik"]).first()
                if not teacher and item["madrasah_email"]:
                    teacher = TeacherProfile.objects.select_related("user").filter(
                        madrasah_email__iexact=item["madrasah_email"]
                    ).first()

                is_new = teacher is None

                if teacher:
                    user = teacher.user
                else:
                    user = CustomUser(role=CustomUser.Role.TEACHER)
                    username_seed = (
                        item["madrasah_email"]
                        or item["nip"]
                        or item["nuptk"]
                        or item["nik"]
                        or item["full_name"]
                    )
                    user.username = _generate_unique_username(username_seed)
                    user.set_password(item["initial_password"])
                    teacher = TeacherProfile(user=user)

                user.full_name = item["full_name"]
                user.email = item["email"]
                user.phone_number = item["phone_number"]
                user.is_school_active = item["is_active"]
                user.role = CustomUser.Role.TEACHER
                user.save()

                teacher.user = user
                teacher.nip = item["nip"] or teacher.nip or None
                teacher.nik = item["nik"] or teacher.nik or None
                teacher.nuptk = item["nuptk"] or teacher.nuptk or None
                teacher.subject = item["subject"]
                teacher.task = item["task"]
                teacher.placement = item["placement"]
                teacher.total_jtm = item["total_jtm"]
                teacher.gender = item["gender"]
                teacher.birth_place = item["birth_place"]
                teacher.birth_date = _parse_iso_date(item["birth_date"])
                teacher.madrasah_email = item["madrasah_email"]
                teacher.employment_status = item["employment_status"]
                teacher.is_active = item["is_active"]
                teacher.save()

            if is_new:
                result["created"] += 1
            else:
                result["updated"] += 1

        except Exception as exc:
            result["failed"] += 1
            result["errors"].append({"sheet": item["sheet"], "row": item["row"], "message": str(exc)})

    return result
