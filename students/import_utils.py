import json
import re
import uuid
from datetime import date, datetime
from pathlib import Path

from django.conf import settings
from django.db import transaction

from accounts.models import CustomUser
from academics.models import AcademicYear, SchoolClass, StudyGroup

from .models import StudentProfile

IMPORT_CACHE_DIR = Path(settings.MEDIA_ROOT) / "import_cache"

HEADER_ALIASES = {
    "nis": "nis",
    "nama": "full_name",
    "nama lengkap": "full_name",
    "nama siswa": "full_name",
    "nama lengkap siswa": "full_name",
    "nisn": "nisn",
    "nik": "nik",
    "tempat lahir": "birth_place",
    "ttl": "birth_date",
    "tgl lahir": "birth_date",
    "tanggal lahir": "birth_date",
    "tingkat rombel": "group_label",
    "kelas": "group_label",
    "kelas rombel": "group_label",
    "rombel": "group_label",
    "umur": "age",
    "status": "status",
    "jk": "gender",
    "jenis kelamin": "gender",
    "alamat": "address",
    "telepon": "phone_number",
    "hp": "phone_number",
    "no hp": "phone_number",
    "no telepon": "phone_number",
    "nomor telepon": "phone_number",
    "nomor hp": "phone_number",
    "kebutuhan khusus": "special_needs",
    "disabilitas": "disability_notes",
    "nomor kip pip": "kip_number",
    "kip pip": "kip_number",
    "ayah": "father_name",
    "nama ayah kandung": "father_name",
    "ibu": "mother_name",
    "nama ibu kandung": "mother_name",
    "wali": "guardian_name",
    "nama wali": "guardian_name",
    "status orang tua": "family_status",
    "status keluarga": "family_status",
    "kondisi orang tua": "family_status",
}


def _normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("\n", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_number_text(value):
    if value is None:
        return ""

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).rstrip("0").rstrip(".")

    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def _parse_date(value):
    if value in (None, ""):
        return None, None

    if isinstance(value, datetime):
        return value.date().isoformat(), None

    if isinstance(value, date):
        return value.isoformat(), None

    text = _clean_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat(), None
        except ValueError:
            continue

    return None, f"Tanggal lahir tidak valid: {text}"


def _parse_iso_date(value):
    if not value:
        return None
    if isinstance(value, date):
        return value
    return date.fromisoformat(value)


def _parse_gender(value):
    text = _clean_text(value).lower()
    if text in {"l", "lk", "laki-laki", "laki laki", "male"}:
        return StudentProfile.Gender.MALE
    if text in {"p", "pr", "perempuan", "female"}:
        return StudentProfile.Gender.FEMALE
    return ""


def _parse_status(value):
    text = _clean_text(value).lower()

    if text == "":
        return True, None
    if text in {"aktif", "active", "ya", "yes", "1", "true"}:
        return True, None
    if text in {"nonaktif", "tidak aktif", "inactive", "no", "0", "false"}:
        return False, None

    return None, f"Status tidak dikenali: {value}"


def _parse_family_status(value):
    text = _clean_text(value).lower()

    if text == "":
        return "", None
    if text in {"lengkap", "orang tua lengkap", "complete"}:
        return StudentProfile.FamilyStatus.COMPLETE, None
    if text in {"yatim", "ayah meninggal"}:
        return StudentProfile.FamilyStatus.ORPHAN_FATHER, None
    if text in {"piatu", "ibu meninggal"}:
        return StudentProfile.FamilyStatus.ORPHAN_MOTHER, None
    if text in {"yatim piatu", "yatim-piatu", "yatim_piatu"}:
        return StudentProfile.FamilyStatus.ORPHAN_BOTH, None
    if text in {"wali", "diasuh wali", "asuh wali"}:
        return StudentProfile.FamilyStatus.UNDER_GUARDIAN, None

    return None, f"Status orang tua tidak dikenali: {value}"


def _split_group_label(label):
    cleaned = _clean_text(label)
    if " - " in cleaned:
        left, right = cleaned.rsplit(" - ", 1)
        return _clean_text(left), _clean_text(right)
    if "-" in cleaned:
        left, right = cleaned.rsplit("-", 1)
        return _clean_text(left), _clean_text(right)
    return cleaned, cleaned


def _infer_level_order(class_name):
    match = re.search(r"(\d+)", class_name or "")
    if match:
        return int(match.group(1))
    return 99


def _infer_entry_year_from_school_class(class_name, active_year):
    if not active_year or not getattr(active_year, "start_date", None):
        return None

    start_year = active_year.start_date.year
    level_order = _infer_level_order(class_name)
    if 7 <= level_order <= 9:
        return start_year - max(level_order - 7, 0)
    return start_year


def _generate_unique_username(preferred):
    base = _clean_text(preferred).lower().replace(" ", "")
    base = re.sub(r"[^a-z0-9._-]", "", base) or "siswa"
    candidate = base
    counter = 1
    while CustomUser.objects.filter(username=candidate).exists():
        candidate = f"{base}-{counter}"
        counter += 1
    return candidate


def _cache_file_path(token):
    IMPORT_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    return IMPORT_CACHE_DIR / f"{token}.json"


def save_import_preview(payload):
    token = uuid.uuid4().hex
    path = _cache_file_path(token)
    path.write_text(json.dumps(payload, ensure_ascii=True), encoding="utf-8")
    return token


def load_import_preview(token):
    path = _cache_file_path(token)
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def delete_import_preview(token):
    path = _cache_file_path(token)
    if path.exists():
        path.unlink()


def _append_error(errors, sheet, row, message):
    errors.append(
        {
            "sheet": sheet,
            "row": row,
            "message": message,
        }
    )


def _resolve_existing_student(nis, nisn, existing_by_nis, existing_by_nisn):
    student_by_nis = existing_by_nis.get(nis) if nis else None
    student_by_nisn = existing_by_nisn.get(nisn) if nisn else None

    if student_by_nis and student_by_nisn and student_by_nis.id != student_by_nisn.id:
        return None, "NIS dan NISN mengarah ke dua siswa yang berbeda."

    return student_by_nisn or student_by_nis, None


def build_student_import_preview(uploaded_file, default_password):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "ok": False,
            "errors": [
                {
                    "sheet": "-",
                    "row": "-",
                    "message": "Paket openpyxl belum terpasang. Jalankan `pip install -r requirements.txt` terlebih dahulu.",
                }
            ],
        }

    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        return {
            "ok": False,
            "errors": [
                {
                    "sheet": "-",
                    "row": "-",
                    "message": "Tahun ajaran aktif belum ditentukan.",
                }
            ],
        }

    workbook = load_workbook(uploaded_file, data_only=True)

    existing_students = list(
        StudentProfile.objects.select_related("user", "study_group").all()
    )
    existing_by_nis = {student.nis: student for student in existing_students if student.nis}
    existing_by_nisn = {student.nisn: student for student in existing_students if student.nisn}

    existing_class_names = set(SchoolClass.objects.values_list("name", flat=True))
    existing_group_names = set(
        StudyGroup.objects.filter(academic_year=active_year).values_list("name", flat=True)
    )

    file_nis_seen = {}
    file_nisn_seen = {}

    rows_payload = []
    errors = []
    classes_to_create = set()
    groups_to_create = set()

    create_count = 0
    update_count = 0

    for sheet in workbook.worksheets:
        if sheet.max_row < 1:
            _append_error(errors, sheet.title, "-", "Sheet kosong.")
            continue

        headers = {}
        for idx, cell in enumerate(sheet[1], start=1):
            normalized = _normalize_header(cell.value)
            if normalized in HEADER_ALIASES:
                headers[HEADER_ALIASES[normalized]] = idx

        if "full_name" not in headers:
            _append_error(errors, sheet.title, 1, "Header wajib tidak ditemukan: Nama Lengkap.")
            continue

        for row_idx in range(2, sheet.max_row + 1):
            row_values = {
                key: sheet.cell(row=row_idx, column=col_idx).value
                for key, col_idx in headers.items()
            }

            full_name = _clean_text(row_values.get("full_name"))
            nis = _normalize_number_text(row_values.get("nis"))
            nisn = _normalize_number_text(row_values.get("nisn"))
            nik = _normalize_number_text(row_values.get("nik"))

            if not full_name and not nis and not nisn and not nik:
                continue

            if not full_name:
                _append_error(errors, sheet.title, row_idx, "Nama lengkap kosong.")
                continue

            if nis:
                prev_row_for_nis = file_nis_seen.get(nis)
                if prev_row_for_nis:
                    _append_error(
                        errors,
                        sheet.title,
                        row_idx,
                        f"NIS {nis} duplikat di file import (baris sebelumnya: {prev_row_for_nis}).",
                    )
                    continue
                file_nis_seen[nis] = row_idx

            if nisn:
                prev_row_for_nisn = file_nisn_seen.get(nisn)
                if prev_row_for_nisn:
                    _append_error(
                        errors,
                        sheet.title,
                        row_idx,
                        f"NISN {nisn} duplikat di file import (baris sebelumnya: {prev_row_for_nisn}).",
                    )
                    continue
                file_nisn_seen[nisn] = row_idx

            group_label = _clean_text(row_values.get("group_label")) or _clean_text(sheet.title)
            school_class_name, study_group_name = _split_group_label(group_label)
            if not school_class_name or not study_group_name:
                _append_error(errors, sheet.title, row_idx, "Format kelas/rombel tidak dikenali.")
                continue

            existing_student, conflict_error = _resolve_existing_student(
                nis=nis,
                nisn=nisn,
                existing_by_nis=existing_by_nis,
                existing_by_nisn=existing_by_nisn,
            )
            if conflict_error:
                _append_error(errors, sheet.title, row_idx, conflict_error)
                continue

            gender = _parse_gender(row_values.get("gender"))
            if not gender:
                _append_error(errors, sheet.title, row_idx, "Jenis kelamin tidak dikenali.")
                continue

            birth_date, birth_date_error = _parse_date(row_values.get("birth_date"))
            if birth_date_error:
                _append_error(errors, sheet.title, row_idx, birth_date_error)
                continue

            is_active, status_error = _parse_status(row_values.get("status"))
            if status_error:
                _append_error(errors, sheet.title, row_idx, status_error)
                continue

            family_status, family_status_error = _parse_family_status(row_values.get("family_status"))
            if family_status_error:
                _append_error(errors, sheet.title, row_idx, family_status_error)
                continue

            action = "update" if existing_student else "create"
            if action == "create":
                create_count += 1
            else:
                update_count += 1

            if school_class_name not in existing_class_names:
                classes_to_create.add(school_class_name)

            if study_group_name not in existing_group_names:
                groups_to_create.add(f"{active_year.id}:{study_group_name}")

            rows_payload.append(
                {
                    "sheet": sheet.title,
                    "row": row_idx,
                    "action": action,
                    "student_id": existing_student.id if existing_student else None,
                    "full_name": full_name,
                    "nis": nis,
                    "nisn": nisn,
                    "nik": nik,
                    "birth_place": _clean_text(row_values.get("birth_place")),
                    "birth_date": birth_date,
                    "gender": gender,
                    "address": _clean_text(row_values.get("address")),
                    "phone_number": _normalize_number_text(row_values.get("phone_number")),
                    "special_needs": _clean_text(row_values.get("special_needs")),
                    "disability_notes": _clean_text(row_values.get("disability_notes")),
                    "kip_number": _normalize_number_text(row_values.get("kip_number")),
                    "father_name": _clean_text(row_values.get("father_name")),
                    "mother_name": _clean_text(row_values.get("mother_name")),
                    "guardian_name": _clean_text(row_values.get("guardian_name")),
                    "family_status": family_status,
                    "school_class_name": school_class_name,
                    "study_group_name": study_group_name,
                    "group_label": group_label,
                    "is_active": is_active,
                    "entry_year": _infer_entry_year_from_school_class(school_class_name, active_year),
                }
            )

    payload = {
        "ok": len(rows_payload) > 0 or len(errors) == 0,
        "active_year_id": active_year.id,
        "active_year_name": active_year.name,
        "default_password": default_password,
        "rows": rows_payload,
        "errors": errors,
        "summary": {
            "total_rows": len(rows_payload) + len(errors),
            "ready_count": len(rows_payload),
            "create_count": create_count,
            "update_count": update_count,
            "error_count": len(errors),
            "new_class_count": len(classes_to_create),
            "new_group_count": len(groups_to_create),
        },
    }
    return payload


def execute_student_import(preview_payload):
    active_year = AcademicYear.objects.get(pk=preview_payload["active_year_id"])

    result = {
        "created": 0,
        "updated": 0,
        "failed": 0,
        "class_created": 0,
        "group_created": 0,
        "errors": [],
    }

    created_classes = set()
    created_groups = set()

    has_nik_field = any(field.name == "nik" for field in StudentProfile._meta.fields)

    for item in preview_payload["rows"]:
        try:
            with transaction.atomic():
                school_class, class_created = SchoolClass.objects.get_or_create(
                    name=item["school_class_name"],
                    defaults={
                        "level_order": _infer_level_order(item["school_class_name"]),
                        "description": "",
                        "is_active": True,
                    },
                )
                if class_created:
                    created_classes.add(school_class.name)

                study_group, group_created = StudyGroup.objects.get_or_create(
                    academic_year=active_year,
                    name=item["study_group_name"],
                    defaults={
                        "school_class": school_class,
                        "capacity": 32,
                        "is_active": True,
                    },
                )
                if group_created:
                    created_groups.add(f"{active_year.id}:{study_group.name}")
                elif study_group.school_class_id != school_class.id:
                    study_group.school_class = school_class
                    study_group.save(update_fields=["school_class"])

                student = None
                if item["student_id"]:
                    student = StudentProfile.objects.select_related("user").filter(
                        pk=item["student_id"]
                    ).first()

                if not student and item["nisn"]:
                    student = StudentProfile.objects.select_related("user").filter(
                        nisn=item["nisn"]
                    ).first()

                if not student and item["nis"]:
                    student = StudentProfile.objects.select_related("user").filter(
                        nis=item["nis"]
                    ).first()

                is_new = student is None

                if student:
                    user = student.user
                    if user is None:
                        user = CustomUser(role=CustomUser.Role.STUDENT)
                        username_seed = item["nis"] or item["nisn"] or item["full_name"]
                        user.username = _generate_unique_username(username_seed)
                        user.set_password(preview_payload["default_password"])
                else:
                    user = CustomUser(role=CustomUser.Role.STUDENT)
                    username_seed = item["nis"] or item["nisn"] or item["full_name"]
                    user.username = _generate_unique_username(username_seed)
                    user.set_password(preview_payload["default_password"])
                    student = StudentProfile(user=user)

                user.full_name = item["full_name"]
                user.phone_number = item["phone_number"]
                user.is_school_active = item["is_active"]
                user.role = CustomUser.Role.STUDENT
                user.save()

                student.user = user
                student.nis = item["nis"] or student.nis or None
                student.nisn = item["nisn"] or None

                if has_nik_field:
                    student.nik = item["nik"] or None

                student.gender = item["gender"]
                student.birth_place = item["birth_place"]
                student.birth_date = _parse_iso_date(item["birth_date"])
                student.address = item["address"]
                student.father_name = item["father_name"]
                student.mother_name = item["mother_name"]
                student.guardian_name = item["guardian_name"]
                student.family_status = item["family_status"]
                student.special_needs = item["special_needs"]
                student.disability_notes = item["disability_notes"]
                student.kip_number = item["kip_number"]
                student.class_name = item["group_label"]
                student.study_group = study_group
                student.entry_year = item["entry_year"]
                student.is_active = item["is_active"]
                student.save()

            if is_new:
                result["created"] += 1
            else:
                result["updated"] += 1

        except Exception as exc:
            result["failed"] += 1
            result["errors"].append(
                {
                    "sheet": item["sheet"],
                    "row": item["row"],
                    "message": str(exc),
                }
            )

    result["class_created"] = len(created_classes)
    result["group_created"] = len(created_groups)
    return result
